import openpyxl
from pathlib import Path
import requests
import urllib.parse
import os,sys,random
from datetime import datetime
from SPARQLWrapper import SPARQLWrapper, JSON
import pystache

url = "https://www.rki.de/DE/Content/InfAZ/N/Neuartiges_Coronavirus/Daten/"
url += "Impfquotenmonitoring.xlsx"
url += ";jsessionid=159D4550C958EDFAA9A49921FA132A35.internet122"
url += "?__blob=publicationFile"
r = requests.get(url, allow_redirects=True)
open('Impfquotenmonitoring.xlsx', 'wb').write(r.content)

xlsx_file = Path(os.getcwd(), 'Impfquotenmonitoring.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj[wb_obj.sheetnames[1]]
sum = 0
for i in range(2,18):
    if type(sheet.cell(row = i, column = 2).value) == int:
        sum += sheet.cell(row = i, column = 2).value

erlaeuterung = wb_obj['Erläuterung']
m_row = erlaeuterung.max_row
for i in range(1, m_row + 1):
    cell_obj = erlaeuterung.cell(row = i, column = 1)
    if "Datenstand: " in str(cell_obj.value):
        datenstand = "Datenstand: "
        if str(erlaeuterung.cell(row = i, column = 2).value) == "=TODAY()":
            datenstand += str(datetime.today().strftime('%d.%m. %Y'))
        datenstand += " " + str(erlaeuterung.cell(row = i, column = 3).value)

endpoint_url = "https://query.wikidata.org/sparql"
query = "SELECT DISTINCT ?city ?cityLabel ?population ?sitelink WHERE {\n"
query +=  "?city wdt:P31/wdt:P279* wd:Q486972;\n"
query +=  "         wdt:P17 wd:Q183;\n"
query +=  "         wdt:P1082 ?population.\n"
query +=  "?sitelink schema:about ?city;\n"
query +=  "          schema:isPartOf <https://de.wikipedia.org/>;\n"
query +=  "FILTER (abs(?population - " + str(sum) + ") < 1000)\n"
query +=  'SERVICE wikibase:label { bd:serviceParam wikibase:language "[AUTO_LANGUAGE],de" }'
query +="\n}"
user_agent = "Impfquotenmonitorvergleich"
sparql = SPARQLWrapper(endpoint_url, agent=user_agent)
sparql.setQuery(query)
sparql.setReturnFormat(JSON)
results = sparql.query().convert()
result = random.choice(results["results"]["bindings"])
city = result["cityLabel"]["value"]
sitelink = result["sitelink"]["value"]
querylink = "https://query.wikidata.org/#" + urllib.parse.quote(query)

renderer = pystache.Renderer()
with open("index.html", "w") as index_file:
    index_file.write(renderer.render_path(
        'index.mustache',
        {
            "sum": f'{sum:,}'.replace(',','.'),
            "city": city,
            "datenstand": datenstand,
            "sitelink": sitelink,
            "querylink": querylink
        }
    ))